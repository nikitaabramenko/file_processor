import yadisk as yd
from tqdm import tqdm
import os
import pandas
import openpyxl as xl


#служебная функция, делает список федеральных округов
def make_region_area_dict(file, group):
    if group == 1 or group == 2:
        header = 2
    elif 'Unnamed: 0' in pandas.read_excel(file, header=6).columns:
        header = 6
    else:
        header = 5
    with pandas.ExcelFile(file) as xlsx:

        sample_file_df = pandas.read_excel(xlsx, header=header)
        areas_list = list(sample_file_df.loc[:, 'Unnamed: 0'])
        result_dict = {}
        for item in areas_list:
            if item.isupper():
                current_key = item
            elif current_key is not None:
                result_dict[item] = current_key
    return result_dict

#служебная функция, добавляет колонку федеральных округов в датафрейм
def make_area_col(row, dict): #new column
    for region, area in dict.items():
        if row['region'].isupper():
            row['area'] = None
        elif row['region'] == region:
            row['area'] = area
    return row


#эта функция возвращает сконкатенированный датафрейм всех таблиц вида 
# СВЕДЕНИЯ О ЛЕЧЕНИИ ЗЛОКАЧЕСТВЕННЫХ НОВООБРАЗОВАНИЙ (ЗНО), ВПЕРВЫЕ
#ЗАРЕГИСТРИРОВАННЫХ В 2021 Г., ПОДЛЕЖАЩИХ РАДИКАЛЬНОМУ ЛЕЧЕНИЮ

def process_group1(file_list): #process group1 func  
    #создаем один раз словарь такого вида {регион: федеральный округ}
    result_dict = make_region_area_dict(file_list[0], group=1)
    print('Processing group 1 files')
    concatenated_df = pandas.DataFrame()
    #main part
    for file in tqdm(file_list):
        workbook_old = xl.load_workbook(filename=file, read_only=True)
        worksheet_old = workbook_old.active
        #adding some identifyers
        year = 2021
        table = worksheet_old['A1'].value.lower().splitlines()[-1].replace(' ', '').rsplit(')')[-1]
        disease  = worksheet_old['A1'].value.lower().splitlines()[-1].replace(' ', '').rsplit(')')[0] + ')'
        ind = 'сведения о лечении злокачественных новообразований (зно), впервые зарегистрированных в 2021 г., подлежащих радикальному лечению'
        code = 'СОП'
        #making dataframe
        df = pandas.read_excel(file, header=2)
        old_cols = df.columns
        new_cols = ['region',
            'число зно, выявленных в отчетном году, радикальное лечение которых закончено в отчетном году',
            'число зно, выявленных в отчетном году, радикальное лечение которых закончено в отчетном году % от впервые выявленных',
            'число зно, выявленных в отчетном году, радикальное лечение которых будет продолжено (не закончено)',
            "число зно, выявленных в отчетном году, радикальное лечение которых будет продолжено (не закончено) % от впервые выявленных",
            "в том числе с использованием методов только хирургического %",
            "в том числе с использованием методов только лучевого %",
            "в том числе с использованием методов только лекарственного %",
            "в том числе с использованием методов комбинир. или компл. (кроме химио-лучевого)%",
            "в том числе с использованием методов химио-лучевого %"
           ]
        #renaming
        df = df.rename(columns={old: new for old, new in zip(old_cols, new_cols)})
        df = df.apply(make_area_col, axis=1, dict=result_dict)
        df.dropna(inplace=True)
        df.reset_index(inplace=True, drop=True)
        df = df[df.columns.tolist()[-1:] + df.columns.tolist()[:-1]]
        df['year'] = year
        df['table'] = table
        df['disease'] = disease
        df['ind'] = ind
        df['code'] = code
        concatenated_df = pandas.concat([concatenated_df, df])
    return concatenated_df
        
#функция возвращает сконкатенированный датафрейм всех таблиц вида 
# Сведения о контингенте больных со злокачественными новообразованиями, 
# состоящем на учете в онкологических учреждениях в 2021 г.

def process_group2(file_list):
    #создаем словарь регион:федеральный округ
    result_dict = make_region_area_dict(file_list[0], group=2)

    concatenated_df = pandas.DataFrame()
    print('Processing group2')
    for file in tqdm(file_list):
        #print(file, 'is being processed')
        workbook_old = xl.load_workbook(filename=file, read_only=True)
        sheetnames = workbook_old.sheetnames
        worksheet_table1 = workbook_old[sheetnames[0]]
        
        year = 2021
        table = worksheet_table1['A1'].value.lower().splitlines()[-1].replace(' ', '').rsplit(')')[-1]
        disease  = worksheet_table1['A1'].value.lower().splitlines()[-1].replace(' ', '').rsplit(')')[0] + ')'
        ind = 'сведения о контингенте больных со злокачественными новообразованиями, состоящем на учете в онкологических учреждениях'
        code = 'СОП'

        if len(sheetnames) > 1:


            with pandas.ExcelFile(file) as xlsx:
                sheet1_df = pandas.read_excel(xlsx, sheetnames[0] ,header=2)
                sheet2_df = pandas.read_excel(xlsx, sheetnames[-1], header=3)

            old_cols_1 = sheet1_df.columns
            old_cols_2 = sheet2_df.columns
            new_cols_2 = [
                'region',
                'Зарегистрировано ЗНО (без учтенных посмертно)',
                "из зарегестрированных диагноз подтвержден морфологически %",
                "из зарегестрированных имели 1 стадию заболевания %",
                "из зарегестрированных имели 2 стадию заболевания %",
                "из зарегестрированных имели 3 стадию заболевания %",
                "из зарегестрированных имели 4 стадию заболевания %",
                "из зарегестрированных стадия заболевания не установлена %",
                "летальность на первом году с момента уст. диагноза %"
            ]
            new_cols_1 = [
                'region',
                'Взято на учет ',
                'в т.ч. выявлены активно %',
                'находились на учете на конец года абсолютное число',
                "находились на учете на конец года на 100тыс населения",
                "из них пять лет и более, абсолютное число",
                "из них пять лет и более, %",
                "индекс накопления контингентов",
                "летальность %"

            ]
            sheet1_df = sheet1_df.rename(columns={old: new for old, new in zip(old_cols_1, new_cols_1)})
            sheet2_df = sheet2_df.rename(columns={old: new for old, new in zip(old_cols_2, new_cols_2)})
            df = sheet1_df.merge(sheet2_df, how='left', on='region')
            df = df.apply(make_area_col, axis=1, dict=result_dict)
            df.dropna(inplace=True)
            df.reset_index(inplace=True, drop=True)
            df = df[df.columns.tolist()[-1:] + df.columns.tolist()[:-1]]
            df['year'] = year
            df['table'] = table
            df['disease'] = disease
            df['ind'] = ind
            df['code'] = code
            concatenated_df = pandas.concat([concatenated_df, df])

        elif len(sheetnames) == 1:

            with pandas.ExcelFile(file) as xlsx:
                sheet1_df = pandas.read_excel(xlsx, sheetnames[0] ,header=2)
                

            old_cols_1 = sheet1_df.columns
            
            
            new_cols_1 = [
                'region',
                'Взято на учет ',
                'в т.ч. выявлены активно %',
                'находились на учете на конец года абсолютное число',
                "находились на учете на конец года на 100тыс населения",
                "из них пять лет и более, абсолютное число",
                "из них пять лет и более, %",
                "индекс накопления контингентов",
                "летальность %"

            ]
            df = sheet1_df.rename(columns={old: new for old, new in zip(old_cols_1, new_cols_1)})
            
            
            df = df.apply(make_area_col, axis=1, dict=result_dict)
            df.dropna(inplace=True)
            df.reset_index(inplace=True, drop=True)
            df = df[df.columns.tolist()[-1:] + df.columns.tolist()[:-1]]
            df['year'] = year
            df['table'] = table
            df['disease'] = disease
            df['ind'] = ind
            df['code'] = code
            concatenated_df = pandas.concat([concatenated_df, df])


    return concatenated_df

#функция возвращает сконкатенированный датафрейм таблиц вида смертность или заболеваемость
def process_group34(file_list, group):
    #print(len(file_list))
    

    concatenated_df = pandas.DataFrame()
    df = pandas.DataFrame()
    print(f'Processing group {group}')
    for file in tqdm(file_list):
        result_dict = make_region_area_dict(file, group=group)
        #print(file, 'processed')
        workbook_old = xl.load_workbook(file, read_only=True)
        worksheet_old= workbook_old.active

        table = worksheet_old['A1'].value.lower()
        ind = worksheet_old['A2'].value.lower()
        code = 'ЗНО'
        if not worksheet_old['A3'].value.replace(' ', '').split(':')[-1].isnumeric():
            year = worksheet_old['B3'].value
        else:
            year = worksheet_old['A3'].value.replace(' ', '').split(':')[-1]
        if  worksheet_old['A4'].value.endswith(':'):
            disease = worksheet_old['B4'].value.lower()
        else:
            disease = worksheet_old['A4'].value.replace(' ', '').split(':')[-1].lower()


        if 'Unnamed: 0' in pandas.read_excel(file, header=6).columns:

            df_raw = pandas.read_excel(file, header=6, index_col='Unnamed: 0')
            old_cols = df_raw.columns
            new_cols = [
                'all',
                'rough',
                'std',
                'error',
                'all men',
                'rough men',
                'std men',
                'error men',
                'all women',
                'rough women',
                'std women',
                'error women'
            ]
            
            df_raw = df_raw.rename(columns={old: new for old, new in zip(old_cols, new_cols)})

            df_all = df_raw.loc[:, 'all':'error']
            df_men = df_raw.loc[:, 'all men':'error men']
            df_women = df_raw.loc[:, 'all women':'error women']
            
            new_cols = [
                'all',
                'rough',
                'std',
                'error'
            ]
            
            df_list = [df_all, df_men, df_women]
            
            
            
            for dataframe, gender in zip(df_list, ['all', 'men', 'women']):
                
                dataframe = dataframe.rename(columns={old:new for old, new in zip(dataframe.columns, new_cols)})
                dataframe['region'] = dataframe.index
                dataframe.reset_index(inplace=True, drop=True)
                dataframe['gender'] = gender
                df = pandas.concat([df, dataframe])
                
                
            df = df.apply(make_area_col, axis=1, dict=result_dict)
            df.dropna(inplace=True, subset=['area'])
            df.reset_index(inplace=True, drop=True)
            
            df['table'] = table
            df['ind'] = ind
            df['year'] = year
            df['code'] = code
            df['disease'] = disease
            
            concatenated_df = pandas.concat([concatenated_df, df])
        elif  'Unnamed: 0' not in pandas.read_excel(file, header=6).columns:

            df = pandas.read_excel(file, header=5, index_col='Unnamed: 0')
            
            old_cols = df.columns
            new_cols = [
                'all',
                'rough',
                'std',
                'error'
            ]
            
            df = df.rename(columns={old: new for old, new in zip(old_cols, new_cols)})
            df['region'] = df.index                
            df = df.apply(make_area_col, axis=1, dict=result_dict)
            df.dropna(inplace=True, subset=['area'])
            df.reset_index(inplace=True, drop=True)
            df['gender'] = 'all'
            df['table'] = table
            df['ind'] = ind
            df['year'] = year
            df['code'] = code
            df['disease'] = disease
            
            concatenated_df = pandas.concat([concatenated_df, df])
        #print(len(concatenated_df))
    return concatenated_df      

# функция делает списки файлов для обработки
def make_file_list(path=None):

    files = os.listdir(path)

    file_g1 = []
    file_g2 = []
    file_g3 = []
    file_g4 = []
    unsorted = []

    for file in tqdm(files):
        #print(f'processing {file}')
        if not file.startswith('~$'):
            try:
                workbook = xl.load_workbook(file, read_only=True)
                worksheet = workbook[workbook.sheetnames[0]]
            except Exception as e:
                print(f'Error processing file {file}: {e}')
                

        try:
            a1 = worksheet['A1'].value.lower().replace(' ', '')
            a2 = worksheet['A2'].value.lower().replace(' ', '')
        except:
            a1 = ''
            a2 = ''
            #print('a1 or a2 doesnt have values')
        if a1.partition('таблица')[0] == '':
            
            if a2 == 'СМЕРТНОСТЬ НАСЕЛЕНИЯ ТЕРРИТОРИЙ РОССИИ ОТ ЗЛОКАЧЕСТВЕННЫХ НОВООБРАЗОВАНИЙ'.lower().replace(' ', '') or a2 == 'смертностьнаселенияроссииотзлокачественныхновообразований':
                file_g4.append(file)
            elif a2 == 'Заболеваемость населения территорий России злокачественными новообразованиями'.lower().replace(' ', ''):
                file_g3.append(file)
            else:
                #print('file cannot be sorted (first if block)')
                unsorted.append(file)
        elif a1.partition('таблица')[0] != '':

            if 'подлежащих радикальному лечению'.replace(' ', '') in a1:
                file_g1.append(file)
            elif 'Сведения о контингенте больных со злокачественными новообразованиями'.lower().replace(' ', '') in a1:
                file_g2.append(file)
            else:
                #print('file cannot be sorted second if block')
                unsorted.append(file)
        else:
            #print('file cannot be sorted (else block)')
            unsorted.append(file)

    return file_g1, file_g2, file_g3, file_g4, unsorted

# закачка файлов с яндекс диска
def download_files_from_disk(app_id, secret_id, token, disk_path, path=None):

    disk = yd.YaDisk(app_id, secret_id, token)

    if path is None:
        print(f'NO PATH ADDED TO DOWNLOAD, CURRENT DIRECTORY {os.getcwd()}')
        print('MAKING `yadisk_files` DIRECTORY')

        try:
            os.mkdir('yadisk_files')
            os.chdir('yadisk_files')
            os.mkdir('preprocessed')
            os.mkdir('processed')
        except Exception as e:
            print(f'Error occured during making directory: {e}')
    else:
        try:
            os.chdir(path)
            os.mkdir('preprocessed')
            os.mkdir('processed')
        except Exception as e:
            print(f'Error occured during making directory: {e}')
        

    os.chdir('preprocessed')


    if disk.check_token():

        list_of_files = []

        for el in list(disk.listdir(f'{disk_path}')):
            if el['path'].endswith('.xlsx'):
                list_of_files.append(el['path'])
        
        print('DOWNLOADING XSLX FILES')
        for file in tqdm(list_of_files):
            disk.download(file.split(':')[1], file.split('/')[-1])
        
        print('DOWNLOADED SUCCESSFULLY')

    else:
        print('FALSE TOKEN, UNABLE TO DOWNLOAD FILES')

    
